#!/usr/bin/env python3
"""
海南离岛免税销售数据爬虫
从海口海关官网抓取月度免税销售数据，解析xlsx/xls文件，保存到data.json

用法:
    python3 scrape.py              # 抓取所有缺失月份
    python3 scrape.py --force      # 强制重新抓取所有月份
    python3 scrape.py --month 2026-02  # 抓取指定月份

添加新月份:
    在 KNOWN_ARTICLES 中添加一行，格式: "YYYY-MM": "xlsx直链URL"
    示例: "2026-04": "http://haikou.customs.gov.cn/.../XXXXXXX/YYYYY.xlsx",
"""

import os
import re
import json
import asyncio
import argparse
import subprocess

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    exit(1)

try:
    import xlrd
except ImportError:
    print("需要 xlrd: pip install xlrd")
    exit(1)

try:
    import nodriver as uc
except ImportError:
    print("需要 nodriver: pip install nodriver")
    exit(1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DATA_FILE = os.path.join(BASE_DIR, "data.json")
LISTING_URL = "https://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/58527f05-{page}.html"
BASE_URL_HTTP = "http://haikou.customs.gov.cn"
BASE_URL_HTTPS = "https://haikou.customs.gov.cn"

# ─── 已知月份 xlsx 直链 ───────────────────────────────────────
# 从海关文章页拿到 xlsx 下载链接后，在这里添加一行。
# xlsx 文件不受 WAF 限制，GitHub Actions 可直接下载。
#
# 获取方法: 浏览器打开文章页 → F12 Network → 点击下载 → 拦截请求 URL
# ─────────────────────────────────────────────────────────────────

KNOWN_ARTICLES = {
    "2024-01": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/5684450/",  # 需要补 xlsx 链接
    "2024-02": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/5757645/",
    "2024-03": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/5825176/",
    "2024-04": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/5884831/",
    "2024-05": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/5941485/",
    "2024-06": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/5992300/",
    "2024-07": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6050921/",
    "2024-08": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6108463/",
    "2024-09": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6159070/",
    "2024-10": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6211777/",
    "2024-11": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6272443/",
    "2024-12": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6326412/",
    "2025-01": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6369960/",
    "2025-02": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6422507/",
    "2025-03": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6471759/",
    "2025-04": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6524076/",
    "2025-05": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6584794/",
    "2025-06": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6630041/",
    "2025-07": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6681393/",
    "2025-08": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6741898/",
    "2025-09": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6782549/",
    "2025-10": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6830442/",
    "2025-11": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6896365/",
    "2025-12": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/6952249/",
    "2026-01": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/7037278/",
    "2026-02": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/7074681/",
    "2026-03": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/7117655/2026042016044569957.xlsx",
    # ↓ 在这里添加新月份 ↓
    # "2026-04": "http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/XXXXXXX/YYYYYYYYYYYYYYYYYYY.xlsx",
}


def _is_xlsx_url(url):
    """判断 URL 是 xlsx 直链还是文章目录路径"""
    return url.endswith(".xlsx") or url.endswith(".xls")


def _get_xlsx_from_article_dir(article_dir):
    """从文章目录路径中发现 xlsx 文件（用 HEAD 探测，不访问 HTML 页面）"""
    # 文章目录路径不以 .xlsx/.xls 结尾，需要从里面找到 xlsx 文件
    # 尝试常见的 xlsx 文件名模式
    if not article_dir.endswith("/"):
        article_dir += "/"

    # 尝试 HEAD 请求探测 xlsx 文件
    # 先试根目录下的常见文件名
    import glob
    import time

    # 无法直接列目录，返回 None
    return None


# ─── curl 工具 ─────────────────────────────────────────────────

def curl_fetch_html(url, referer=None):
    """用 curl + 浏览器 UA 抓取 HTML"""
    cmd = [
        "curl", "-s", "-k", "-L", "--max-time", "60",
        "-w", "\n__CURL_HTTP_CODE__%{http_code}",
        "-H", "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
        "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "-H", "Accept-Language: zh-CN,zh;q=0.9,en;q=0.8",
    ]
    if referer:
        cmd += ["-H", f"Referer: {referer}"]
    cmd.append(url)
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=65)
        raw = result.stdout.decode("utf-8", errors="ignore") if result.stdout else ""
        http_code = 0
        m = re.search(r"__CURL_HTTP_CODE__(\d+)", raw)
        if m:
            http_code = int(m.group(1))
            raw = raw[:m.start()]
        if result.returncode != 0:
            return ""
        if http_code and http_code >= 400:
            return ""
        if len(raw) < 500:
            return ""
        return raw
    except Exception:
        return ""


def curl_download(url, output_path, referer=None):
    """用 curl 下载文件"""
    cmd = [
        "curl", "-s", "-k", "-L",
        "-o", output_path, "--max-time", "60",
        "-H", "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
              "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
    ]
    if referer:
        cmd += ["-H", f"Referer: {referer}"]
    cmd.append(url)
    try:
        result = subprocess.run(cmd, capture_output=True, timeout=65)
        return result.returncode == 0
    except Exception:
        return False


def _extract_xlsx_from_html(html):
    """从 HTML 中提取 xlsx/xls 下载链接"""
    m = re.search(r'href="([^"]+\.xls[x]?)"', html)
    if m:
        url = m.group(1)
        if not url.startswith("http"):
            url = BASE_URL_HTTPS + url
        return url
    return None


# ─── nodriver 浏览器管理（列表页发现用）─────────────────────────

_browser = None


async def get_browser():
    """获取全局 nodriver 浏览器实例"""
    global _browser
    if _browser is None:
        chrome_path = os.environ.get("CHROME_PATH") or _find_chrome()
        kwargs = dict(
            headless=True,
            sandbox=False,
            browser_args=[
                "--no-sandbox", "--disable-dev-shm-usage",
                "--ignore-certificate-errors", "--disable-gpu",
            ],
        )
        if chrome_path:
            kwargs["browser_executable_path"] = chrome_path
        print(f"  启动浏览器: {chrome_path or '自动检测'}")
        _browser = await uc.start(**kwargs)
    return _browser


def _find_chrome():
    import shutil
    for name in ["google-chrome-stable", "google-chrome", "chromium", "chrome"]:
        found = shutil.which(name)
        if found:
            return found
    for p in ["/opt/hostedtoolcache/setup-chrome/chromium/stable/x64/chrome",
              "/usr/bin/google-chrome-stable", "/usr/bin/google-chrome"]:
        if os.path.exists(p):
            return p
    return None


async def close_browser():
    global _browser
    if _browser:
        try:
            _browser.stop()
        except Exception:
            pass
        finally:
            _browser = None


async def nw_fetch_html(url, wait_sec=10):
    browser = await get_browser()
    try:
        page = await browser.get(url)
        await asyncio.sleep(wait_sec)
        html = await page.get_content()
        if "504" in html and "连接超时" in html:
            return ""
        if "502" in html and "Bad Gateway" in html:
            return ""
        return html
    except Exception:
        return ""


# ─── 列表页发现（仅本机可用）───────────────────────────────────

async def discover_articles_from_listing():
    """从列表页发现新文章（本机可用，GitHub Actions 被 WAF 拦截）"""
    articles = {}
    no_match_streak = 0
    for page_num in range(1, 10):
        urls = [
            f"http://haikou.customs.gov.cn/haikou_customs/605737/fdzdgknr82/605745/58527f05-{page_num}.html",
            LISTING_URL.format(page=page_num),
        ]
        html = ""
        for url in urls:
            html = curl_fetch_html(url)
            if html:
                break
        if not html:
            html = await nw_fetch_html(LISTING_URL.format(page=page_num), wait_sec=12)
        if not html:
            break

        pattern = r'<a[^>]*href="([^"]+)"[^>]*>([^<]*(?:离岛免税|免税销售|免税购物)[^<]*)</a>'
        matches = re.findall(pattern, html)
        if not matches:
            no_match_streak += 1
            if no_match_streak >= 2:
                break
            continue

        no_match_streak = 0
        for href, title in matches:
            m = re.search(r"(\d{4})年(\d{1,2})月", title)
            if m:
                key = f"{m.group(1)}-{m.group(2).zfill(2)}"
                if key not in articles:
                    articles[key] = href

        await asyncio.sleep(1)

    return articles


# ─── 解析 ──────────────────────────────────────────────────────

def parse_xlsx(filepath):
    """解析 xlsx/xls 文件，提取免税数据"""
    rows = []
    if filepath.endswith(".xlsx"):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.worksheets[0]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            rows.append([str(c) if c is not None else "" for c in row])
    else:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        for r in range(ws.nrows):
            rows.append([
                str(ws.cell_value(r, c)) if ws.cell_value(r, c) != "" else ""
                for c in range(ws.ncols)
            ])

    month_data = {}
    for row in rows:
        row_str = " ".join(row)

        if "免税购物金额" in row_str and "实际人次" not in row_str and "件数" not in row_str:
            unit = row[2].strip() if len(row) > 2 else ""
            multiplier = 10000 if "亿" in unit else 1
            nums = _extract_nums(row)
            if len(nums) >= 1: month_data["amount"] = round(nums[0] * multiplier, 2)
            if len(nums) >= 2: month_data["amount_yoy"] = round(nums[1], 2)
            if len(nums) >= 3: month_data["amount_cumulative"] = round(nums[2] * multiplier, 2)
            if len(nums) >= 4: month_data["amount_cumulative_yoy"] = round(nums[3], 2)

        elif "免税购物实际人次" in row_str:
            nums = _extract_nums(row)
            if len(nums) >= 1: month_data["visitors"] = round(nums[0], 4)
            if len(nums) >= 2: month_data["visitors_yoy"] = round(nums[1], 2)
            if len(nums) >= 3: month_data["visitors_cumulative"] = round(nums[2], 4)
            if len(nums) >= 4: month_data["visitors_cumulative_yoy"] = round(nums[3], 2)

        elif "免税购物件数" in row_str:
            nums = _extract_nums(row)
            if len(nums) >= 1: month_data["items"] = round(nums[0], 4)
            if len(nums) >= 2: month_data["items_yoy"] = round(nums[1], 2)
            if len(nums) >= 3: month_data["items_cumulative"] = round(nums[2], 4)
            if len(nums) >= 4: month_data["items_cumulative_yoy"] = round(nums[3], 2)

    if month_data.get("amount") and month_data.get("visitors") and month_data["visitors"] > 0:
        month_data["per_capita"] = round(month_data["amount"] / month_data["visitors"], 2)
    if month_data.get("amount_cumulative") and month_data.get("visitors_cumulative") and month_data.get("visitors_cumulative", 0) > 0:
        month_data["per_capita_cumulative"] = round(month_data["amount_cumulative"] / month_data["visitors_cumulative"], 2)

    return month_data


def _extract_nums(row):
    nums = []
    for c in row:
        try:
            nums.append(float(c))
        except (ValueError, TypeError):
            pass
    return nums


# ─── 数据存储 ──────────────────────────────────────────────────

def load_existing_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            raw = json.load(f)
        if isinstance(raw, list):
            return raw
        if isinstance(raw, dict) and "data" in raw:
            return raw["data"]
    return []


def save_data(data):
    data.sort(key=lambda x: x["month"])
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已保存 {len(data)} 条记录到 data.json")


# ─── 抓取逻辑 ──────────────────────────────────────────────────

async def scrape_month(month_key, url_or_path, force=False):
    """抓取单个月份的数据"""
    existing = load_existing_data()
    existing_months = {d["month"] for d in existing}

    if not force and month_key in existing_months:
        print(f"  {month_key}: 已存在，跳过")
        return True

    # 判断是 xlsx 直链还是文章目录路径
    if _is_xlsx_url(url_or_path):
        xlsx_url = url_or_path
        print(f"  {month_key}: 使用 xlsx 直链")
    else:
        # 文章目录路径，需要从文章页获取 xlsx 链接
        # 注意: GitHub Actions 上会被 WAF 拦截
        print(f"  {month_key}: 需要从文章页获取 xlsx 链接...")
        print(f"  ⚠️  GitHub Actions 上无法访问文章页，请添加 xlsx 直链")
        return False

    # 下载
    ext = ".xlsx" if xlsx_url.endswith(".xlsx") else ".xls"
    tmp_file = os.path.join(DATA_DIR, f"{month_key}{ext}")

    print(f"  {month_key}: 下载中...")
    if not curl_download(xlsx_url, tmp_file):
        print(f"  {month_key}: 下载失败")
        return False

    # 解析
    print(f"  {month_key}: 解析中...")
    data = parse_xlsx(tmp_file)
    if not data:
        print(f"  {month_key}: 解析失败")
        return False

    data["month"] = month_key
    existing = [d for d in existing if d["month"] != month_key]
    existing.append(data)
    save_data(existing)

    print(f"  {month_key}: ✅ 完成 (金额={data.get('amount')}万元, 人次={data.get('visitors')}万, 件数={data.get('items')}万)")
    return True


async def scrape_all(force=False):
    """抓取所有已知月份"""
    os.makedirs(DATA_DIR, exist_ok=True)

    print(f"已知 {len(KNOWN_ARTICLES)} 个月份")
    print()

    success = 0
    fail = 0
    for month_key in sorted(KNOWN_ARTICLES.keys()):
        if await scrape_month(month_key, KNOWN_ARTICLES[month_key], force=force):
            success += 1
        else:
            fail += 1
        await asyncio.sleep(0.5)

    print(f"\n完成: 成功 {success}, 失败 {fail}")


async def async_main(args):
    try:
        if args.month:
            os.makedirs(DATA_DIR, exist_ok=True)
            if args.month in KNOWN_ARTICLES:
                await scrape_month(args.month, KNOWN_ARTICLES[args.month], force=True)
            else:
                print(f"未知月份 {args.month}")
                print("已知月份:", ", ".join(sorted(KNOWN_ARTICLES.keys())))
        else:
            await scrape_all(force=args.force)
    finally:
        await close_browser()


def main():
    parser = argparse.ArgumentParser(description="海南离岛免税销售数据爬虫")
    parser.add_argument("--force", action="store_true", help="强制重新抓取所有月份")
    parser.add_argument("--month", type=str, help="只抓取指定月份 (格式: 2026-02)")
    args = parser.parse_args()

    import gc
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(async_main(args))
    finally:
        gc.collect()
        try:
            loop.run_until_complete(asyncio.sleep(0.1))
        except RuntimeError:
            pass
        loop.close()


if __name__ == "__main__":
    main()
